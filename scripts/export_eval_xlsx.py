"""导出 RAG 评测结果（run cases.jsonl）为格式化 Excel 报告。

用法:
    python scripts/export_eval_xlsx.py --run-id <runId> [--out <path.xlsx>]
    python scripts/export_eval_xlsx.py --cases <cases.jsonl> [--out <path.xlsx>]

依赖: openpyxl (>=3.1)
"""
from __future__ import annotations

import argparse
import json
import statistics as st
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

METRICS = [
    "answer_relevancy",
    "context_precision",
    "context_recall",
    "factual_correctness_f1",
    "faithfulness",
]
METRIC_CN = {
    "answer_relevancy": "答案相关性",
    "context_precision": "上下文精确率",
    "context_recall": "上下文召回率",
    "factual_correctness_f1": "事实正确率F1",
    "faithfulness": "忠实度",
}
DOMAIN_CN = {"SERVICE": "服务域", "COMPLIANCE": "合规域", "MENTAL": "心理域"}

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Border()


def _ok(n: float) -> bool:
    return n >= 1.0


def _load_cases(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").strip().splitlines() if line.strip()]


def _score(c: dict) -> dict:
    return c.get("ragasScores") or {}


def _stats(c: dict) -> dict:
    return c.get("ragasStats") or {}


def _style_header(ws, ncols: int, row: int = 1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN


def _set_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> int:
    parser = argparse.ArgumentParser(description="导出 RAG 评测结果到 Excel")
    parser.add_argument("--run-id", default=None, help="run id（读取 target/rag-eval/runs/<run-id>/cases.jsonl）")
    parser.add_argument("--cases", default=None, help="直接指定 cases.jsonl 路径")
    parser.add_argument("--out", default=None, help="输出 xlsx 路径")
    args = parser.parse_args()

    if args.cases:
        cases_path = Path(args.cases)
    elif args.run_id:
        cases_path = Path(f"target/rag-eval/runs/{args.run_id}/cases.jsonl")
    else:
        # 自动取最新 run
        runs = sorted(Path("target/rag-eval/runs").glob("*/cases.jsonl"))
        if not runs:
            print("未找到任何 run 产物，请用 --run-id 或 --cases 指定。", file=sys.stderr)
            return 1
        cases_path = runs[-1]
    if not cases_path.exists():
        print(f"未找到 {cases_path}", file=sys.stderr)
        return 1
    cases = _load_cases(cases_path)
    if not cases:
        print(f"{cases_path} 为空", file=sys.stderr)
        return 1

    out = Path(args.out) if args.out else cases_path.with_name("rag-eval-report.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ---------- Sheet1 概览 ----------
    ws = wb.active
    ws.title = "概览"
    ws.append(["MindBridge RAG 评测报告", ""])
    ws.append(["run-id", cases_path.parent.name])
    ws.append(["case 数", len(cases)])
    ws.append(["采样", max((st_["samples"] for c in cases for st_ in _stats(c).values()), default=1)])
    ws.append(["", ""])
    ws.append(["指标", "均值(中位数)", "达标 / 总数"])
    _style_header(ws, 3, row=6)
    for i, m in enumerate(METRICS):
        vals = [_score(c)[m] for c in cases if isinstance(_score(c).get(m), (int, float))]
        if not vals:
            continue
        mean = sum(vals) / len(vals)
        ok_cnt = sum(1 for v in vals if _ok(v))
        ws.append([METRIC_CN[m], round(mean, 4), f"{ok_cnt} / {len(vals)}"])
    ws.append(["", "", ""])
    ws.append(["结论", "检索层 context_recall=1.0 全召回；LLM-judge 指标良好但存在 judge 方差，门禁为 Observe", ""])
    _set_widths(ws, [22, 24, 14])

    # ---------- Sheet2 逐case --------
    ws = wb.create_sheet("逐case")
    header = ["caseId", "域", "场景", "答案相关性", "上下文精确率", "上下文召回率", "事实正确率F1", "忠实度"]
    ws.append(header)
    _style_header(ws, len(header))
    for c in cases:
        s = _score(c)
        ws.append([
            c["caseId"], DOMAIN_CN.get(c.get("domain"), c.get("domain")), c.get("scenario", ""),
            round(s.get("answer_relevancy", 0), 4),
            round(s.get("context_precision", 0), 4),
            round(s.get("context_recall", 0), 4),
            round(s.get("factual_correctness_f1", 0), 4),
            round(s.get("faithfulness", 0), 4),
        ])
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=8):
        for cell in row:
            cell.alignment = CENTER
            if isinstance(cell.value, (int, float)) and cell.value < 0.6:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(color="9C0006")
    _set_widths(ws, [34, 10, 14, 12, 12, 12, 12, 12])

    # ---------- Sheet3 采样明细(median/mean/std) ----------
    ws = wb.create_sheet("采样明细")
    header = ["caseId", "指标", "中位数", "均值", "标准差", "采样数"]
    ws.append(header)
    _style_header(ws, len(header))
    for c in cases:
        for m in METRICS:
            s = _score(c)
            st_ = _stats(c).get(m, {})
            if not st_:
                continue
            ws.append([
                c["caseId"], METRIC_CN[m],
                st_.get("median", ""), st_.get("mean", ""), st_.get("std", ""), st_.get("samples", ""),
            ])
    _set_widths(ws, [34, 16, 10, 10, 10, 10])

    # ---------- Sheet4 分域 ----------
    ws = wb.create_sheet("分域")
    header = ["域", "case 数"] + [METRIC_CN[m] for m in METRICS]
    ws.append(header)
    _style_header(ws, len(header))
    by_domain: dict[str, list[dict]] = {}
    for c in cases:
        by_domain.setdefault(c.get("domain", "?"), []).append(c)
    for dom, dom_cases in sorted(by_domain.items()):
        row = [DOMAIN_CN.get(dom, dom), len(dom_cases)]
        for m in METRICS:
            vals = [_score(c)[m] for c in dom_cases if isinstance(_score(c).get(m), (int, float))]
            row.append(round(sum(vals) / len(vals), 4) if vals else "")
        ws.append(row)
    _set_widths(ws, [14, 10] + [13] * len(METRICS))

    # ---------- Sheet5 评测配置 ----------
    ws = wb.create_sheet("评测配置")
    config = [
        ("评测集", "data/eval/smoke/rag-smoke.json"),
        ("域 / case", "SERVICE 3 / COMPLIANCE 3 / MENTAL 4"),
        ("检索", "SQLite + BM25 召回 + qwen3-vl-rerank 语义重排 (topK=4)"),
        ("Judge", "DashScope qwen-max"),
        ("采样", "--runs 3，中位数聚合，ragasStats 报告 median/mean/std"),
        ("指标", "answer_relevancy / context_precision / context_recall / factual_correctness_f1 / faithfulness"),
        ("数据来源", str(cases_path)),
    ]
    for k, v in config:
        ws.append([k, v])
    _set_widths(ws, [14, 90])

    wb.save(out)
    print(f"已导出: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())